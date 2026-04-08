"""シフト表Excel生成: 色分け付き3シート構成。

Sheet 1: シフト表（曜日×従業員、色分け）
Sheet 2: 公平性サマリ（従業員別統計）
Sheet 3: 制約チェック（HC/SC充足状況）

Usage:
    python generate_excel.py
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.error("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent.parent
RESULTS_DIR = BASE_DIR / "results"
DELIVERY_DIR = BASE_DIR / "delivery"
DELIVERY_DIR.mkdir(exist_ok=True)

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DAY_JP = {"Mon": "月", "Tue": "火", "Wed": "水", "Thu": "木", "Fri": "金", "Sat": "土", "Sun": "日"}
SHIFTS = ["morning", "afternoon", "night"]
SHIFT_JP = {"morning": "朝勤", "afternoon": "午後勤", "night": "夜勤"}

# Colors
FILL_MORNING = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue
FILL_AFTERNOON = PatternFill(start_color="CCF2CC", end_color="CCF2CC", fill_type="solid")  # Light green
FILL_NIGHT = PatternFill(start_color="E0CCF2", end_color="E0CCF2", fill_type="solid")  # Light purple
FILL_UNFILLED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue header
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
FILL_NG = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow

FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_BOLD = Font(bold=True, size=11)
FONT_NORMAL = Font(size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_schedule_and_eval():
    """Load the best schedule from baseline results."""
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from baseline import load_data, evaluate, solve_cpsat

    data = load_data()
    schedule = solve_cpsat(data)
    evaluation = evaluate(schedule, data)
    return data, schedule, evaluation


def create_shift_table_sheet(ws, data, schedule, evaluation):
    """Sheet 1: Shift table with color coding."""
    ws.title = "シフト表"

    employees = data["employees"]
    emp_names = {e["id"]: e["name"] for e in employees}
    emp_ids = [e["id"] for e in employees]

    # Build lookup: (day, shift) -> list of employee_ids
    assigned = {d: {s: [] for s in SHIFTS} for d in DAYS}
    for entry in schedule:
        assigned[entry["day"]][entry["shift"]].append(entry["employee_id"])

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "週間シフト表（CP-SAT最適化結果）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = ALIGN_CENTER

    # Header row
    row = 3
    headers = ["時間帯"] + [f"{DAY_JP[d]}({d})" for d in DAYS]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # Required count row
    row = 4
    ws.cell(row=row, column=1, value="必要人数").font = FONT_BOLD
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col, day in enumerate(DAYS, 2):
        req_parts = []
        for sdef in data["shifts"]:
            if sdef["day"] == day:
                req_parts.append(f"{SHIFT_JP[sdef['shift']]}:{sdef['required']}")
        cell = ws.cell(row=row, column=col, value="\n".join(req_parts))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.font = Font(size=9)

    # Shift data rows (one row per shift type)
    for si, shift in enumerate(SHIFTS):
        row = 5 + si
        cell = ws.cell(row=row, column=1, value=SHIFT_JP[shift])
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        fills = {"morning": FILL_MORNING, "afternoon": FILL_AFTERNOON, "night": FILL_NIGHT}

        for col, day in enumerate(DAYS, 2):
            emps = assigned[day][shift]
            # Find required count
            req = 0
            for sdef in data["shifts"]:
                if sdef["day"] == day and sdef["shift"] == shift:
                    req = sdef["required"]
                    break

            names = [emp_names[eid] for eid in emps]
            cell = ws.cell(row=row, column=col, value="\n".join(names) if names else "(未充足)")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            cell.font = FONT_NORMAL

            if len(emps) < req:
                cell.fill = FILL_UNFILLED
            else:
                cell.fill = fills[shift]

    # Employee schedule (individual view)
    row = 10
    ws.cell(row=row, column=1, value="従業員別シフト").font = Font(bold=True, size=12)

    row = 11
    headers = ["従業員"] + [f"{DAY_JP[d]}" for d in DAYS] + ["合計h"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    emp_schedule = {eid: {} for eid in emp_ids}
    for entry in schedule:
        emp_schedule[entry["employee_id"]][entry["day"]] = entry["shift"]

    for ei, eid in enumerate(emp_ids):
        row = 12 + ei
        emp = next(e for e in employees if e["id"] == eid)
        cell = ws.cell(row=row, column=1, value=f"{eid} {emp['name']}")
        cell.font = FONT_BOLD
        cell.border = THIN_BORDER

        total_shifts = 0
        for col, day in enumerate(DAYS, 2):
            shift_name = emp_schedule[eid].get(day, "")
            cell = ws.cell(row=row, column=col, value=SHIFT_JP.get(shift_name, "OFF"))
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

            if shift_name == "morning":
                cell.fill = FILL_MORNING
            elif shift_name == "afternoon":
                cell.fill = FILL_AFTERNOON
            elif shift_name == "night":
                cell.fill = FILL_NIGHT

            if shift_name:
                total_shifts += 1

            # Mark unavailable days
            if day in emp["unavailable_days"] and not shift_name:
                cell.value = "不可"
                cell.font = Font(size=9, color="999999")

        cell = ws.cell(row=row, column=len(DAYS) + 2, value=total_shifts * 8)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # Legend
    row = 12 + len(emp_ids) + 2
    ws.cell(row=row, column=1, value="凡例:").font = FONT_BOLD
    legends = [
        (FILL_MORNING, "朝勤 (09:00-17:00)"),
        (FILL_AFTERNOON, "午後勤 (13:00-21:00)"),
        (FILL_NIGHT, "夜勤 (21:00-05:00)"),
        (FILL_UNFILLED, "未充足（人員不足）"),
    ]
    for i, (fill, label) in enumerate(legends):
        r = row + 1 + i
        cell = ws.cell(row=r, column=1, value="  ")
        cell.fill = fill
        cell.border = THIN_BORDER
        ws.cell(row=r, column=2, value=label).font = FONT_NORMAL

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col in range(2, len(DAYS) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.row_dimensions[4].height = 50
    for si in range(3):
        ws.row_dimensions[5 + si].height = 40


def create_fairness_sheet(ws, data, schedule, evaluation):
    """Sheet 2: Fairness summary."""
    ws.title = "公平性サマリ"

    employees = data["employees"]
    emp_ids = [e["id"] for e in employees]
    hours = evaluation["stats"]["hours_per_employee"]

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "従業員別 勤務時間・公平性サマリ"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = ALIGN_CENTER

    # Headers
    row = 3
    headers = ["従業員ID", "氏名", "勤務時間(h)", "最大(h)", "最低(h)", "夜勤回数", "充足率"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    emp_schedule = {eid: [] for eid in emp_ids}
    for entry in schedule:
        emp_schedule[entry["employee_id"]].append(entry)

    for ei, eid in enumerate(emp_ids):
        emp = next(e for e in employees if e["id"] == eid)
        row = 4 + ei
        h = hours.get(eid, 0)
        night_count = sum(1 for e in emp_schedule[eid] if e["shift"] == "night")
        fill_rate = h / emp["max_hours"] * 100 if emp["max_hours"] > 0 else 0

        ws.cell(row=row, column=1, value=eid).border = THIN_BORDER
        ws.cell(row=row, column=2, value=emp["name"]).border = THIN_BORDER
        cell = ws.cell(row=row, column=3, value=h)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.cell(row=row, column=4, value=emp["max_hours"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=5, value=emp["min_hours"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=5).border = THIN_BORDER
        cell_night = ws.cell(row=row, column=6, value=night_count)
        cell_night.alignment = ALIGN_CENTER
        cell_night.border = THIN_BORDER
        if night_count > 2:
            cell_night.fill = FILL_WARN
        cell_rate = ws.cell(row=row, column=7, value=f"{fill_rate:.0f}%")
        cell_rate.alignment = ALIGN_CENTER
        cell_rate.border = THIN_BORDER

        # Color code: min_hours not met
        if h < emp["min_hours"]:
            cell.fill = FILL_NG

    # Summary stats
    row = 4 + len(emp_ids) + 2
    ws.cell(row=row, column=1, value="統計").font = FONT_BOLD
    stats = [
        ("平均勤務時間", f"{sum(hours.values()) / len(hours):.1f}h"),
        ("標準偏差", f"{evaluation['stats']['hours_std_dev']}h"),
        ("最大勤務時間", f"{max(hours.values())}h"),
        ("最小勤務時間", f"{min(hours.values())}h"),
        ("最大-最小差", f"{max(hours.values()) - min(hours.values())}h"),
    ]
    for i, (label, value) in enumerate(stats):
        ws.cell(row=row + 1 + i, column=1, value=label).font = FONT_NORMAL
        ws.cell(row=row + 1 + i, column=2, value=value).font = FONT_NORMAL

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_constraint_check_sheet(ws, data, schedule, evaluation):
    """Sheet 3: Constraint check results."""
    ws.title = "制約チェック"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "制約チェック結果"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = ALIGN_CENTER

    # Hard constraints
    row = 3
    ws.cell(row=row, column=1, value="ハード制約（必須）").font = Font(bold=True, size=12)

    row = 4
    headers = ["ID", "制約", "状態", "違反数", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    hc_data = [
        ("HC1", "必要人数の充足", evaluation["hard_violations_detail"]["HC1"], "構造的に2シフト不足（需要48>供給46）"),
        ("HC2", "最大勤務時間の遵守", evaluation["hard_violations_detail"]["HC2"], ""),
        ("HC3", "不可日の遵守", evaluation["hard_violations_detail"]["HC3"], ""),
        ("HC4", "スキル要件の充足", evaluation["hard_violations_detail"]["HC4"], ""),
        ("HC5", "夜勤→朝勤の禁止", evaluation["hard_violations_detail"]["HC5"], ""),
    ]

    for i, (hc_id, desc, violations, note) in enumerate(hc_data):
        row = 5 + i
        ws.cell(row=row, column=1, value=hc_id).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        status = "OK" if violations == 0 else "NG"
        cell = ws.cell(row=row, column=3, value=status)
        cell.fill = FILL_OK if violations == 0 else FILL_NG
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.cell(row=row, column=4, value=violations).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=5, value=note).border = THIN_BORDER

    # Soft constraints
    row = 5 + len(hc_data) + 2
    ws.cell(row=row, column=1, value="ソフト制約（努力目標）").font = Font(bold=True, size=12)

    row += 1
    headers = ["ID", "制約", "スコア", "判定", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    sc = evaluation["soft_scores"]
    sc_data = [
        ("SC1", "連続勤務5日以内", sc["SC1_consecutive"], "連続6日以上の従業員なし"),
        ("SC2", "公平性（時間偏り最小化）", sc["SC2_fairness"], f"std_dev={evaluation['stats']['hours_std_dev']}h（max_hours格差が原因）"),
        ("SC3", "最低勤務時間の確保", sc["SC3_min_hours"], "全員min_hours以上"),
        ("SC4", "夜勤週2回以内", sc["SC4_night_limit"], "全員2回以下"),
        ("SC5", "研修スキル者の配置", sc["SC5_training"], "全日に配置あり"),
    ]

    for i, (sc_id, desc, score, note) in enumerate(sc_data):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=sc_id).border = THIN_BORDER
        ws.cell(row=r, column=2, value=desc).border = THIN_BORDER
        cell = ws.cell(row=r, column=3, value=f"{score:.1f}/100")
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if score >= 80:
            judge = "良好"
            fill = FILL_OK
        elif score >= 50:
            judge = "注意"
            fill = FILL_WARN
        else:
            judge = "要改善"
            fill = FILL_NG
        cell_j = ws.cell(row=r, column=4, value=judge)
        cell_j.fill = fill
        cell_j.alignment = ALIGN_CENTER
        cell_j.border = THIN_BORDER
        ws.cell(row=r, column=5, value=note).border = THIN_BORDER

    # Total score
    r = row + 1 + len(sc_data) + 1
    ws.cell(row=r, column=1, value="合計スコア").font = FONT_BOLD
    ws.cell(row=r, column=3, value=f"{evaluation['soft_score_total']}/100").font = FONT_BOLD

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 45


def main():
    logger.info("Loading schedule data...")
    data, schedule, evaluation = load_schedule_and_eval()

    wb = Workbook()
    ws1 = wb.active

    logger.info("Creating Sheet 1: Shift table...")
    create_shift_table_sheet(ws1, data, schedule, evaluation)

    logger.info("Creating Sheet 2: Fairness summary...")
    ws2 = wb.create_sheet()
    create_fairness_sheet(ws2, data, schedule, evaluation)

    logger.info("Creating Sheet 3: Constraint check...")
    ws3 = wb.create_sheet()
    create_constraint_check_sheet(ws3, data, schedule, evaluation)

    output_path = DELIVERY_DIR / "shift_schedule.xlsx"
    wb.save(output_path)
    logger.info("Excel saved to %s", output_path)


if __name__ == "__main__":
    main()

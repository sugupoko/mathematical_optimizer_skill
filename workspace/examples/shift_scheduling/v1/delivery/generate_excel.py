"""
シフト表 Excel 生成スクリプト

受け入れ要件に基づく:
- 行=従業員、列=曜日、セル=シフト名
- 色分け: 朝=青、昼=緑、夜=紫、未充足=赤背景
- サマリ: 各人の週合計時間、夜勤回数、公平性指標
- 制約チェック結果シート付き
"""
import csv
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
RESULTS_DIR = BASE_DIR / "results"
DELIVERY_DIR = BASE_DIR / "delivery"

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DAY_JP = {"Mon": "月", "Tue": "火", "Wed": "水", "Thu": "木", "Fri": "金", "Sat": "土", "Sun": "日"}
SHIFT_NAMES = ["morning", "afternoon", "night"]
SHIFT_JP = {"morning": "朝勤", "afternoon": "昼勤", "night": "夜勤"}
SHIFT_HOURS = 8

# 色定義
FILL_MORNING = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # 薄い青
FILL_AFTERNOON = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")  # 薄い緑
FILL_NIGHT = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")  # 薄い紫
FILL_SHORTAGE = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")  # 薄い赤
FILL_HEADER = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
FILL_WARNING = PatternFill(start_color="F5B041", end_color="F5B041", fill_type="solid")  # オレンジ
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_BOLD = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_SMALL = Font(size=9, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def load_data():
    employees = []
    with open(DATA_DIR / "employees.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            row["skills"] = set(row["skills"].split(","))
            row["max_hours"] = int(row["max_hours_per_week"])
            row["min_hours"] = int(row["min_hours_per_week"])
            row["unavailable"] = set(row["unavailable_days"].split(",")) - {""}
            employees.append(row)

    shifts = []
    with open(DATA_DIR / "shifts.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            row["required_count"] = int(row["required_count"])
            shifts.append(row)

    return employees, shifts


def load_solver_result():
    """ベースライン結果からソルバー解を復元（improve の +1人案が採用されていない前提）"""
    # baseline.py を再実行して assignment を取得する代わりに、
    # baseline スクリプトを import して使う
    import importlib.util
    spec = importlib.util.spec_from_file_location("baseline", BASE_DIR / "scripts" / "baseline.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    employees, shifts = mod.load_data()
    assignment, status, solver = mod.baseline_solver(employees, shifts)
    return assignment, employees, shifts


def build_schedule_matrix(assignment, employees, shifts):
    """assignment dict → 人×日のシフト名マトリクス"""
    n_emp = len(employees)
    matrix = [[None] * 7 for _ in range(n_emp)]

    for (e_idx, s_idx), val in assignment.items():
        if val:
            day_idx = s_idx // 3
            shift_type = SHIFT_NAMES[s_idx % 3]
            matrix[e_idx][day_idx] = shift_type

    return matrix


def check_constraints(matrix, employees, shifts):
    """手修正後にも使える制約チェック。違反リストを返す"""
    violations = []
    n_emp = len(employees)

    for e_idx, emp in enumerate(employees):
        hours = sum(SHIFT_HOURS for d in range(7) if matrix[e_idx][d] is not None)
        night_count = sum(1 for d in range(7) if matrix[e_idx][d] == "night")
        name = emp["name"]

        # HC2: 最大勤務時間
        if hours > emp["max_hours"]:
            violations.append(("HC2", f"{name}: 週{hours}h（上限{emp['max_hours']}h）", "エラー"))

        # HC3: 利用不可日
        for d in range(7):
            if matrix[e_idx][d] is not None and DAYS[d] in emp["unavailable"]:
                violations.append(("HC3", f"{name}: {DAY_JP[DAYS[d]]}曜は出勤不可", "エラー"))

        # HC4: スキル
        for d in range(7):
            if matrix[e_idx][d] is not None:
                s_idx = d * 3 + SHIFT_NAMES.index(matrix[e_idx][d])
                req_skill = shifts[s_idx]["required_skills"]
                if req_skill not in emp["skills"]:
                    violations.append(("HC4", f"{name}: {DAY_JP[DAYS[d]]}{SHIFT_JP[matrix[e_idx][d]]}に必要なスキル「{req_skill}」なし", "エラー"))

        # HC5: 夜勤翌日朝勤
        for d in range(6):
            if matrix[e_idx][d] == "night" and matrix[e_idx][d + 1] == "morning":
                violations.append(("HC5", f"{name}: {DAY_JP[DAYS[d]]}夜勤→{DAY_JP[DAYS[d+1]]}朝勤（休息不足）", "エラー"))

        # SC1: 連続勤務
        consecutive = 0
        max_consecutive = 0
        for d in range(7):
            if matrix[e_idx][d] is not None:
                consecutive += 1
                max_consecutive = max(max_consecutive, consecutive)
            else:
                consecutive = 0
        if max_consecutive > 5:
            violations.append(("SC1", f"{name}: {max_consecutive}日連続勤務（推奨5日以下）", "注意"))

        # SC3: 最低勤務時間
        if hours < emp["min_hours"]:
            violations.append(("SC3", f"{name}: 週{hours}h（最低{emp['min_hours']}h 未達）", "注意"))

        # SC4: 夜勤回数
        if night_count > 2:
            violations.append(("SC4", f"{name}: 夜勤{night_count}回（推奨2回以下）", "注意"))

    # HC1: 各シフトの人員
    for d in range(7):
        for s_name_idx, s_name in enumerate(SHIFT_NAMES):
            s_idx = d * 3 + s_name_idx
            required = shifts[s_idx]["required_count"]
            actual = sum(1 for e in range(n_emp) if matrix[e][d] == s_name)
            if actual < required:
                violations.append(("HC1", f"{DAY_JP[DAYS[d]]}{SHIFT_JP[s_name]}: {actual}/{required}名（{required - actual}名不足）", "エラー"))

    return violations


def create_shift_sheet(wb, matrix, employees, shifts):
    """メインのシフト表シート"""
    ws = wb.active
    ws.title = "シフト表"

    # タイトル
    ws.merge_cells("A1:I1")
    ws["A1"] = "週間シフト表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # ヘッダー行
    headers = ["従業員"] + [f"{DAY_JP[d]}（{d}）" for d in DAYS] + ["週合計"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # データ行
    for e_idx, emp in enumerate(employees):
        row = e_idx + 4
        # 名前
        cell = ws.cell(row=row, column=1, value=emp["name"])
        cell.font = FONT_BOLD
        cell.border = THIN_BORDER

        hours = 0
        for d in range(7):
            col = d + 2
            shift = matrix[e_idx][d]
            cell = ws.cell(row=row, column=col)
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

            if shift is None:
                cell.value = "休"
                cell.font = FONT_SMALL
            else:
                cell.value = SHIFT_JP[shift]
                cell.font = FONT_NORMAL
                hours += SHIFT_HOURS
                if shift == "morning":
                    cell.fill = FILL_MORNING
                elif shift == "afternoon":
                    cell.fill = FILL_AFTERNOON
                elif shift == "night":
                    cell.fill = FILL_NIGHT

        # 週合計
        cell = ws.cell(row=row, column=9, value=f"{hours}h/{emp['max_hours']}h")
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # シフト人数サマリ行
    row_summary_start = len(employees) + 5
    ws.cell(row=row_summary_start, column=1, value="").border = THIN_BORDER

    for s_name_idx, s_name in enumerate(SHIFT_NAMES):
        row = row_summary_start + s_name_idx
        cell = ws.cell(row=row, column=1, value=f"  {SHIFT_JP[s_name]}人数")
        cell.font = FONT_SMALL
        cell.fill = FILL_SUBHEADER
        cell.border = THIN_BORDER

        for d in range(7):
            col = d + 2
            s_idx = d * 3 + s_name_idx
            required = shifts[s_idx]["required_count"]
            actual = sum(1 for e in range(len(employees)) if matrix[e][d] == s_name)
            cell = ws.cell(row=row, column=col, value=f"{actual}/{required}")
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            if actual < required:
                cell.fill = FILL_SHORTAGE
                cell.font = Font(bold=True, size=10, color="C0392B")

    # 列幅調整
    ws.column_dimensions["A"].width = 16
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["I"].width = 14


def create_fairness_sheet(wb, matrix, employees):
    """公平性サマリシート"""
    ws = wb.create_sheet("公平性サマリ")

    headers = ["従業員", "週勤務時間", "上限", "稼働率", "夜勤回数", "連続勤務(最大)", "状態"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    total_hours = []
    for e_idx, emp in enumerate(employees):
        row = e_idx + 2
        hours = sum(SHIFT_HOURS for d in range(7) if matrix[e_idx][d] is not None)
        night_count = sum(1 for d in range(7) if matrix[e_idx][d] == "night")
        total_hours.append(hours)

        # 連続勤務
        consecutive = max_consecutive = 0
        for d in range(7):
            if matrix[e_idx][d] is not None:
                consecutive += 1
                max_consecutive = max(max_consecutive, consecutive)
            else:
                consecutive = 0

        utilization = hours / emp["max_hours"] * 100

        # 状態判定
        status = "OK"
        if hours < emp["min_hours"]:
            status = "⚠ 最低時間未達"
        elif night_count > 2:
            status = "⚠ 夜勤過多"
        elif max_consecutive > 5:
            status = "⚠ 連続勤務超過"

        values = [
            emp["name"], f"{hours}h", f"{emp['max_hours']}h",
            f"{utilization:.0f}%", f"{night_count}回", f"{max_consecutive}日", status
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            if status.startswith("⚠"):
                ws.cell(row=row, column=7).fill = FILL_WARNING

    # 統計行
    import statistics
    row = len(employees) + 3
    avg = statistics.mean(total_hours)
    sd = statistics.stdev(total_hours) if len(total_hours) > 1 else 0
    ws.cell(row=row, column=1, value="統計").font = FONT_BOLD
    ws.cell(row=row, column=2, value=f"平均 {avg:.1f}h / 標準偏差 {sd:.1f}h").font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    # 列幅
    widths = [16, 14, 10, 10, 12, 16, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def create_check_sheet(wb, violations):
    """制約チェック結果シート"""
    ws = wb.create_sheet("制約チェック")

    headers = ["制約ID", "内容", "レベル"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    if not violations:
        ws.cell(row=2, column=1, value="✓ 全ての制約を満たしています").font = Font(color="27AE60", bold=True, size=11)
        ws.merge_cells("A2:C2")
    else:
        for i, (cid, desc, level) in enumerate(violations):
            row = i + 2
            ws.cell(row=row, column=1, value=cid).font = FONT_NORMAL
            ws.cell(row=row, column=2, value=desc).font = FONT_NORMAL
            ws.cell(row=row, column=3, value=level).font = FONT_NORMAL
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = ALIGN_CENTER
            if level == "エラー":
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = FILL_SHORTAGE
            elif level == "注意":
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = FILL_WARNING

    # 凡例
    row = max(3, len(violations) + 3)
    ws.cell(row=row, column=1, value="凡例:").font = FONT_BOLD
    ws.cell(row=row + 1, column=1, value="エラー").fill = FILL_SHORTAGE
    ws.cell(row=row + 1, column=2, value="ハード制約違反（必ず修正が必要）").font = FONT_SMALL
    ws.cell(row=row + 2, column=1, value="注意").fill = FILL_WARNING
    ws.cell(row=row + 2, column=2, value="ソフト制約違反（可能なら修正）").font = FONT_SMALL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10


def main():
    assignment, employees, shifts = load_solver_result()
    matrix = build_schedule_matrix(assignment, employees, shifts)
    violations = check_constraints(matrix, employees, shifts)

    wb = Workbook()
    create_shift_sheet(wb, matrix, employees, shifts)
    create_fairness_sheet(wb, matrix, employees)
    create_check_sheet(wb, violations)

    output_path = DELIVERY_DIR / "shift_schedule.xlsx"
    wb.save(output_path)
    print(f"シフト表を生成しました: {output_path}")
    print(f"  シート1: シフト表（色分け付き）")
    print(f"  シート2: 公平性サマリ（勤務時間・夜勤回数・稼働率）")
    print(f"  シート3: 制約チェック（{len(violations)}件の指摘）")

    # 違反があれば表示
    if violations:
        print(f"\n制約チェック結果:")
        for cid, desc, level in violations:
            print(f"  [{level}] {cid}: {desc}")


if __name__ == "__main__":
    main()
